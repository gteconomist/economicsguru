#!/usr/bin/env python3
"""
Fetch the New York Fed Quarterly Report on Household Debt and Credit and
refresh the three CSV baselines built from it.

Why this exists
---------------
These three CSVs used to be hand-seeded, which meant they only advanced when
someone remembered to re-upload them. By 2026-08-12 the balance and
delinquency panels were two quarters behind (2026Q1 released 2026-05-12 and
2026Q2 released 2026-08-11 were both sitting unclaimed) while the pages that
render them looked perfectly healthy, because the monthly series alongside
them kept updating.

The NY Fed publishes the full report data as a real .xlsx at a stable,
predictable URL, and -- unlike the licence-restricted NAR series on FRED --
EVERY quarterly file carries the complete history back to 2003Q1. So this
is a true replacement for the hand-seeding, not a trailing-window patch:
a rebuild from any single file reproduces the entire series.

Source
------
  https://www.newyorkfed.org/medialibrary/interactives/householdcredit/data/xls/HHD_C_Report_YYYYQn.xlsx
  Index of all quarters: https://www.newyorkfed.org/microeconomics/hhdc/background.html

  Published ~6 weeks after quarter end, simultaneously with the report itself.
  Public Federal Reserve material, free to republish.
  Attribution: "Source: New York Fed Consumer Credit Panel/Equifax".

Outputs
-------
  data/historical/nyfed_household_debt.csv    quarter, credit_card, home_equity,
                                              auto, student, other  ($ trillions)
  data/historical/nyfed_delinquency.csv       quarter, credit_card, mortgage,
                                              auto, student  (% of balance 90+ dpd)
  data/historical/ny_fed_hhdc_mortgage.csv    quarter_end, mortgage_debt_t
                                              ($ trillions)

Consumed by fetch_consumer.py (debt + delinquency charts) and
fetch_housing_mortgage_activity.py (mortgage debt outstanding), so this must
run BEFORE both of them in refresh.yml.

Design notes
------------
* Sheets are located by the CONTENT of their header row, not by page number.
  The NY Fed reorders and renumbers pages between reports -- "Page 3 Data"
  today is not guaranteed to be the balance table next year, and a silent
  mismatch would write the wrong numbers into the right columns. Matching on
  the loan-type labels fails loudly instead.
* Idempotent: rewrites a CSV only when a value actually changes.
* Non-blocking: any failure leaves the committed CSVs untouched and exits 0,
  so the rest of the refresh proceeds on the previous baselines. Staleness is
  scripts/check_freshness.py's job to report.
"""

import csv
import io
import sys
import datetime as dt
from pathlib import Path
from urllib import request, error

try:
    import openpyxl
except ImportError:
    print("FETCH FAILED: openpyxl is not installed "
          "(add `pip install openpyxl` to the workflow)", file=sys.stderr)
    sys.exit(0)

REPO_ROOT      = Path(__file__).resolve().parents[1]
HISTORICAL_DIR = REPO_ROOT / "data" / "historical"

DEBT_CSV       = HISTORICAL_DIR / "nyfed_household_debt.csv"
DELINQ_CSV     = HISTORICAL_DIR / "nyfed_delinquency.csv"
MORTGAGE_CSV   = HISTORICAL_DIR / "ny_fed_hhdc_mortgage.csv"

BASE = ("https://www.newyorkfed.org/medialibrary/interactives/"
        "householdcredit/data/xls/")

UA = ("Mozilla/5.0 (compatible; economicsguru.com data refresh; "
      "+https://economicsguru.com/about/)")

# Header labels that identify each table. Compared case-insensitively as a
# subset, so an added column (the NY Fed has added "Other" splits before)
# doesn't break the match.
BALANCE_LABELS = {"mortgage", "he revolving", "auto loan",
                  "credit card", "student loan", "other"}
DELINQ_LABELS  = {"mortgage", "heloc", "auto", "cc", "student loan", "other"}

# Header labels alone are NOT enough to identify the delinquency table: three
# consecutive sheets ("Percent of Balance 90+ Days Delinquent", "New Delinquent
# Balances", "New Seriously Delinquent Balances") carry an identical loan-type
# header. The sheet TITLE in row 1 is what separates them, so each table is
# matched on title AND header.
BALANCE_TITLE = "total debt balance and its composition"
DELINQ_TITLE  = "percent of balance 90+ days delinquent"

# workbook column label -> our CSV column
BALANCE_MAP = {
    "credit card":  "credit_card",
    "he revolving": "home_equity",
    "auto loan":    "auto",
    "student loan": "student",
    "other":        "other",
}
DELINQ_MAP = {
    "cc":           "credit_card",
    "mortgage":     "mortgage",
    "auto":         "auto",
    "student loan": "student",
}


# ------------------------------------------------------------------ helpers

def _quarter_candidates(today=None):
    """Newest-first list of (year, quarter) worth trying, covering ~1 year.

    The report lands ~6 weeks after quarter end, so the current quarter is
    never available and the prior one may not be either. Walking backwards
    and taking the first URL that exists handles both the normal case and a
    delayed release without hardcoding a calendar.
    """
    today = today or dt.date.today()
    y, q = today.year, (today.month - 1) // 3 + 1
    out = []
    for _ in range(5):
        q -= 1
        if q == 0:
            q, y = 4, y - 1
        out.append((y, q))
    return out


def _fetch_workbook_bytes():
    """Download the newest available quarterly workbook.

    Returns (bytes, "YYYYQn"). Raises if nothing in the candidate window
    could be fetched.
    """
    errors = []
    for y, q in _quarter_candidates():
        # The served filename is mixed-case; the lowercase path also resolves,
        # but try the canonical spelling first.
        for stem in (f"HHD_C_Report_{y}Q{q}.xlsx", f"hhd_c_report_{y}q{q}.xlsx"):
            url = BASE + stem
            try:
                req = request.Request(url, headers={"User-Agent": UA})
                with request.urlopen(req, timeout=90) as r:
                    blob = r.read()
                if not blob[:2] == b"PK":
                    raise RuntimeError(
                        f"{url} did not return an xlsx (first bytes {blob[:8]!r})")
                print(f"  fetched {stem} ({len(blob):,} bytes)", file=sys.stderr)
                return blob, f"{y}Q{q}"
            except error.HTTPError as e:
                errors.append(f"{stem}: HTTP {e.code}")
            except Exception as e:                      # noqa: BLE001
                errors.append(f"{stem}: {e}")
    raise RuntimeError("no NY Fed workbook could be fetched -- tried: "
                       + "; ".join(errors))


def _norm(cell):
    return str(cell).strip().lower() if cell is not None else ""


def _parse_quarter(label):
    """'26:Q2' -> '2026Q2'. Returns None if the cell isn't a quarter label."""
    s = str(label or "").strip()
    if ":" not in s:
        return None
    yy, q = s.split(":", 1)
    q = q.strip().upper()
    if not (yy.strip().isdigit() and q.startswith("Q") and q[1:].isdigit()):
        return None
    yy = int(yy.strip())
    year = 2000 + yy if yy < 50 else 1900 + yy
    return f"{year}Q{int(q[1:])}"


def _find_table(wb, want_labels, want_title):
    """Return {quarter: {label: value}} for the sheet matching `want_title`
    (substring of row 1) whose header row carries `want_labels`.

    Scans the first six rows of each "* Data" sheet for a header row, so a
    layout shift of a row or two doesn't break it. Raises if no sheet matches
    or if more than one does -- a silent wrong-sheet match would write
    plausible-looking but incorrect numbers.
    """
    matches = []
    for name in wb.sheetnames:
        if not name.strip().lower().endswith("data"):
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(max_row=200, max_col=12, values_only=True))
        title = _norm(rows[0][0]) if rows and rows[0] else ""
        if want_title not in title:
            continue
        for hdr_i, row in enumerate(rows[:6]):
            labels = {_norm(c) for c in row[1:] if _norm(c)}
            if not want_labels.issubset(labels):
                continue
            header = [_norm(c) for c in row]
            table = {}
            for r in rows[hdr_i + 1:]:
                quarter = _parse_quarter(r[0])
                if not quarter:
                    continue
                vals = {}
                for i in range(1, min(len(header), len(r))):
                    if header[i] and isinstance(r[i], (int, float)):
                        vals[header[i]] = float(r[i])
                if vals:
                    table[quarter] = vals
            if table:
                matches.append((name, table))
            break
    if not matches:
        raise RuntimeError(
            f"no sheet titled ~{want_title!r} with header labels "
            f"{sorted(want_labels)} -- the NY Fed workbook layout has changed")
    if len(matches) > 1:
        raise RuntimeError(
            f"ambiguous: {len(matches)} sheets match {want_title!r} + "
            f"{sorted(want_labels)} "
            f"({', '.join(n for n, _ in matches)})")
    name, table = matches[0]
    print(f"  matched sheet {name!r} ({len(table)} quarters)", file=sys.stderr)
    return table


def _fmt(v, decimals=6):
    s = f"%.{decimals}f" % v
    return s.rstrip("0").rstrip(".") if "." in s else s


def _quarter_end(quarter):
    """'2026Q2' -> '2026-06-30'."""
    y, q = quarter.split("Q")
    month = int(q) * 3
    day = 30 if month in (6, 9) else 31
    return f"{y}-{month:02d}-{day:02d}"


def _write_csv(path, header, rows_by_key, key_name):
    """Rewrite `path` if any value changed. Returns True if written.

    Preserves the file's existing line endings -- these CSVs are CRLF, and
    silently flipping them turns a three-line data update into a whole-file
    diff.
    """
    newline = "\r\n"
    merged = {}
    if path.exists():
        raw = path.read_bytes()
        if b"\r\n" not in raw:
            newline = "\n"
        existing = list(csv.reader(io.StringIO(raw.decode("utf-8"))))
        if existing:
            old_header = [c.strip() for c in existing[0]]
            for r in existing[1:]:
                if not r or not r[0].strip():
                    continue
                merged[r[0].strip()] = dict(zip(old_header, [c.strip() for c in r]))

    for key, vals in rows_by_key.items():
        row = merged.get(key, {})
        row[key_name] = key
        row.update(vals)
        merged[key] = row

    body = []
    for key in sorted(merged):
        row = merged[key]
        body.append([str(row.get(col, "")).strip() for col in header])

    out = newline.join([",".join(header)] + [",".join(r) for r in body]) + newline
    if path.exists() and path.read_bytes().decode("utf-8") == out:
        return False
    path.write_bytes(out.encode("utf-8"))
    return True


# --------------------------------------------------------------------- main

def main():
    print("Fetching NY Fed Household Debt and Credit workbook...", file=sys.stderr)
    blob, quarter = _fetch_workbook_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)

    balances = _find_table(wb, BALANCE_LABELS, BALANCE_TITLE)
    delinq   = _find_table(wb, DELINQ_LABELS, DELINQ_TITLE)

    if quarter not in balances:
        raise RuntimeError(
            f"workbook is for {quarter} but its balance table stops at "
            f"{max(balances) if balances else 'nothing'}")

    debt_rows = {
        q: {ours: _fmt(vals[theirs])
            for theirs, ours in BALANCE_MAP.items() if theirs in vals}
        for q, vals in balances.items()
    }
    delinq_rows = {
        q: {ours: _fmt(vals[theirs])
            for theirs, ours in DELINQ_MAP.items() if theirs in vals}
        for q, vals in delinq.items()
    }
    mortgage_rows = {
        _quarter_end(q): {"mortgage_debt_t": "%.3f" % vals["mortgage"]}
        for q, vals in balances.items() if "mortgage" in vals
    }

    wrote = []
    for path, header, rows, key in (
        (DEBT_CSV, ["quarter", "credit_card", "home_equity", "auto", "student", "other"],
         debt_rows, "quarter"),
        (DELINQ_CSV, ["quarter", "credit_card", "mortgage", "auto", "student"],
         delinq_rows, "quarter"),
        (MORTGAGE_CSV, ["quarter_end", "mortgage_debt_t"],
         mortgage_rows, "quarter_end"),
    ):
        if _write_csv(path, header, rows, key):
            wrote.append(path.name)

    print(f"  latest quarter in workbook: {max(balances)}", file=sys.stderr)
    print(f"  CSVs updated: {', '.join(wrote) if wrote else 'none (already current)'}",
          file=sys.stderr)


if __name__ == "__main__":
    try:
        main()
    except (error.URLError, RuntimeError) as e:
        print(f"FETCH FAILED: {e}", file=sys.stderr)
