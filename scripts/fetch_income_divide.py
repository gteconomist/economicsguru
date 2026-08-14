#!/usr/bin/env python3
"""
Build the Income Divide ("K-shaped economy") page from the New York Fed's
Economic Heterogeneity Indicators (EHIs).

Why this exists
---------------
The charts everyone circulates on the K-shaped consumer -- spending of lower-
and higher-income households relative to middle-income households -- come from
Bank of America's internal card data. That is proprietary: it ships only as
chart images inside a monthly PDF, with no data file, no API, and a BofA Global
Research copyright. It cannot be republished and it cannot be automated.

The NY Fed EHIs are the public equivalent. They are built on a 200,000-household
panel from Numerator and published as a real .xlsx at a stable URL, free to
republish as Federal Reserve material. This script turns that workbook into the
CSV baselines and JSON payload behind /consumer/income-divide/.

IMPORTANT -- this is not the BofA chart with a different label
---------------------------------------------------------------
Built from the EHIs, the higher-income/middle-income spending ratio was still
RISING through 2026-04 while the lower-income ratio had only recovered to about
parity. In other words the public data says the K is still widening, which is
the opposite of the "K is waning" read from the card data. Different panel
(survey vs cards), different universe (all retail vs card-discretionary), and
the EHI ends four months earlier. Do not caption these charts as agreeing with
the sell-side version.

Also note BofA's "discretionary" cut (total ex gas, groceries, utilities) is NOT
reproducible here: the EHIs publish retail, gas, and food-and-beverage as three
separate indices with no expenditure weights, so they cannot be netted out.

Source
------
  https://www.newyorkfed.org/medialibrary/research/interactives/data/
      economic-heterogeneity-indicators/downloads/ehi-national-data.xlsx

  Released QUARTERLY -- February, May, and September -- carrying monthly data
  with roughly a one-month lag at release. As of 2026-08 the file holds data
  through 2026-04. The May->September gap is four months, not three, which is
  why the freshness tolerance for this dataset is wider than the site's other
  quarterly series. See scripts/check_freshness.py.

  Public Federal Reserve material. Citation required by the workbook:
    "Federal Reserve Bank of New York, Economic Heterogeneity Indicators,
     https://www.newyorkfed.org/research/economic-heterogeneity-indicators"
  The EHIs are explicitly not official estimates of the New York Fed, the
  Federal Reserve System, or the FOMC -- the page carries that disclaimer.

Outputs
-------
  data/historical/nyfed_ehi_spending.csv    month, {low,mid,high}_retail,
                                            {low,mid,high}_foodbev,
                                            top_125_175, top_175_225,
                                            top_225_250, top_250p
  data/historical/nyfed_ehi_gaps.csv        month, infl_headline, infl_bottom40,
                                            infl_mid40, infl_top20,
                                            edu_urate_gap
  data/income_divide.json                   payload for EG.boot()

Design notes
------------
* Sheets are located by NAME and then VALIDATED against the column labels we
  expect. The NY Fed has reordered sheets between releases; a silent mismatch
  would write high-income numbers into the low-income column, which is exactly
  the failure mode that is impossible to spot on a finished chart.
* Everything is derived from the workbook's "cumulative growth" INDEX LEVELS,
  never from its "year-over-year" columns. The two families do not reconcile
  (the yoy column is not the yoy of the published index), and an index level is
  the only thing a ratio can honestly be built from. YoY here is computed from
  the index and labelled as such on the page.
* The high-income sub-brackets ($125-175k etc.) exist ONLY on the 2023 base, so
  those series legitimately start in 2023. Blank before that is correct.
* "Inflation by Demographic" holds GAPS in percentage points, not levels: the
  columns sit within +/-0.6 and roughly population-weight to zero, while
  "Inflation by Category" -> Headline carries the actual rate (9.06% in
  2022-06). The page therefore reconstructs each group's inflation as
  headline + gap, and says so on the chart. Do not plot the gap columns as
  though they were rates.
* The workbook's "Earnings Ratios (real)" -> College Ratio column (~54-57) is
  deliberately NOT published here: the sheet documents no units and the value
  could be either direction of the college/no-college ratio. A series whose
  definition we are guessing at does not belong on the site. The College
  Unemployment Gap is unambiguous (percentage points, college minus
  non-college) and carries the education story instead.
* Idempotent: rewrites a CSV only when a value actually changes.
* Non-blocking: any failure leaves the committed CSVs and JSON untouched and
  exits 0, so the rest of the refresh proceeds. Staleness is
  scripts/check_freshness.py's job to report.
"""

import csv
import io
import json
import sys
import datetime as dt
from pathlib import Path
from urllib import request, error

try:
    import openpyxl
except ImportError:
    print("FETCH FAILED: openpyxl is not installed "
          "(add `pip install openpyxl` to the workflow)", file=sys.stderr)
    sys.exit(0)

REPO_ROOT      = Path(__file__).resolve().parents[1]
HISTORICAL_DIR = REPO_ROOT / "data" / "historical"
SPEND_CSV      = HISTORICAL_DIR / "nyfed_ehi_spending.csv"
GAPS_CSV       = HISTORICAL_DIR / "nyfed_ehi_gaps.csv"
OUT_PATH       = REPO_ROOT / "data" / "income_divide.json"

URL = ("https://www.newyorkfed.org/medialibrary/research/interactives/data/"
       "economic-heterogeneity-indicators/downloads/ehi-national-data.xlsx")

UA = ("Mozilla/5.0 (compatible; economicsguru.com data refresh; "
      "+https://economicsguru.com/about/)")

# ---------------------------------------------------------------- sheet specs
#
# label -> the exact column header in the workbook. Validated on load; a missing
# label aborts the whole run rather than writing a half-built payload.

RETAIL_SHEET  = "Retail Spending (real)"
FOODBEV_SHEET = "FoodBev Spending (real)"
INFL_SHEET    = "Inflation by Demographic"
INFLCAT_SHEET = "Inflation by Category"
URATE_SHEET   = "URate Gaps"

TIER_2020 = {
    "low":  "Low Income (<$40k), cumulative growth (2020)",
    "mid":  "Middle Income ($40k-$125k), cumulative growth (2020)",
    "high": "High Income ($125k+), cumulative growth (2020)",
}

TOP_2023 = {
    "top_125_175": "$125k-$175k, cumulative growth (2023)",
    "top_175_225": "$175k-$225k, cumulative growth (2023)",
    "top_225_250": "$225k-$250k, cumulative growth (2023)",
    "top_250p":    "$250k+, cumulative growth (2023)",
}

INFL_COLS = {
    "infl_bottom40": "Bottom 40%",
    "infl_mid40":    "40% - 80%",
    "infl_top20":    "Top 20%",
}

HEADLINE_COL = "Headline"
URATE_COL    = "College Unemployment Gap"


# ---------------------------------------------------------------- fetch/parse

def download(url):
    req = request.Request(url, headers={"User-Agent": UA})
    with request.urlopen(req, timeout=120) as resp:
        return resp.read()


def sheet_table(wb, sheet_name):
    """Return (headers, rows) for a Date-indexed EHI sheet.

    headers: {column label -> column index}
    rows:    list of (date, [cell values]) with the date already normalised to
             a 'YYYY-MM' string. Rows without a real date are dropped -- the
             workbook pads every sheet with blank rows and a title block.
    """
    if sheet_name not in wb.sheetnames:
        raise RuntimeError("sheet %r not found; workbook has %r"
                           % (sheet_name, wb.sheetnames))
    ws = wb[sheet_name]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]

    hdr_i = None
    for i, row in enumerate(grid):
        if row and isinstance(row[0], str) and row[0].strip() == "Date":
            hdr_i = i
            break
    if hdr_i is None:
        raise RuntimeError("no 'Date' header row in sheet %r" % sheet_name)

    headers = {}
    for j, cell in enumerate(grid[hdr_i]):
        if isinstance(cell, str) and cell.strip():
            headers[cell.strip()] = j

    rows = []
    for row in grid[hdr_i + 1:]:
        if not row:
            continue
        d = row[0]
        if isinstance(d, dt.datetime):
            d = d.date()
        if not isinstance(d, dt.date):
            continue
        rows.append(("%04d-%02d" % (d.year, d.month), row))
    return headers, rows


def require(headers, wanted, sheet_name):
    missing = [c for c in wanted if c not in headers]
    if missing:
        raise RuntimeError("sheet %r is missing expected column(s) %r -- the NY "
                           "Fed has changed the workbook layout; the parser must "
                           "be re-checked before this data is published"
                           % (sheet_name, missing))


def col(rows, headers, label):
    """[(month, value_or_None)] for one column."""
    j = headers[label]
    out = []
    for month, row in rows:
        v = row[j] if j < len(row) else None
        out.append((month, float(v) if isinstance(v, (int, float)) else None))
    return out


# ---------------------------------------------------------------- transforms

def to_pairs(series, decimals=3):
    return [[m, (None if v is None else round(v, decimals))] for m, v in series]


def ratio_3m(num, den):
    """3-month average of num/den, aligned on month. Mirrors the construction
    used on the sell-side card-data charts (ratio of index levels, 3m average)
    so the two are visually comparable even though the panels differ."""
    dmap = dict(den)
    raw = []
    for m, n in num:
        d = dmap.get(m)
        raw.append((m, None if (n is None or d in (None, 0)) else n / d))
    out, buf = [], []
    for m, v in raw:
        buf.append(v)
        if len(buf) > 3:
            buf.pop(0)
        if len(buf) == 3 and all(x is not None for x in buf):
            out.append((m, sum(buf) / 3.0))
        else:
            out.append((m, None))
    return out


def yoy_from_index(series):
    """Percent change vs 12 months earlier, computed off the index level."""
    vmap = dict(series)
    out = []
    for m, v in series:
        y, mo = int(m[:4]), int(m[5:])
        prev = vmap.get("%04d-%02d" % (y - 1, mo))
        out.append((m, None if (v is None or prev in (None, 0))
                    else (v / prev - 1.0) * 100.0))
    return out


def latest(series):
    for m, v in reversed(series):
        if v is not None:
            return m, v
    return None, None


def kpi(series, decimals=2, lag=1):
    """{value, delta, label} against `lag` observations earlier."""
    pts = [(m, v) for m, v in series if v is not None]
    if not pts:
        return {"value": None, "delta": None, "label": None}
    m, v = pts[-1]
    prev = pts[-1 - lag][1] if len(pts) > lag else None
    return {"value": round(v, decimals),
            "delta": None if prev is None else round(v - prev, decimals),
            "label": m}


# ---------------------------------------------------------------- CSV writing

def write_csv(path, header, rows):
    """Write only if content changed. Returns True when the file was rewritten."""
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(header)
    for r in rows:
        w.writerow(["" if x is None else x for x in r])
    new = buf.getvalue()
    if path.exists() and path.read_text() == new:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(new)
    return True


# ---------------------------------------------------------------------- main

def main():
    print("Fetching NY Fed EHI national workbook ...", flush=True)
    raw = download(URL)
    print("  %d bytes" % len(raw), flush=True)
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)

    # ---- retail spending -------------------------------------------------
    r_h, r_rows = sheet_table(wb, RETAIL_SHEET)
    require(r_h, list(TIER_2020.values()) + list(TOP_2023.values()), RETAIL_SHEET)
    retail = {k: col(r_rows, r_h, lbl) for k, lbl in TIER_2020.items()}
    tops   = {k: col(r_rows, r_h, lbl) for k, lbl in TOP_2023.items()}

    # ---- food & beverage spending (the staples counterweight) ------------
    f_h, f_rows = sheet_table(wb, FOODBEV_SHEET)
    require(f_h, list(TIER_2020.values()), FOODBEV_SHEET)
    foodbev = {k: col(f_rows, f_h, lbl) for k, lbl in TIER_2020.items()}

    # ---- inflation by income quintile group ------------------------------
    i_h, i_rows = sheet_table(wb, INFL_SHEET)
    require(i_h, list(INFL_COLS.values()), INFL_SHEET)
    infl = {k: col(i_rows, i_h, lbl) for k, lbl in INFL_COLS.items()}

    c_h, c_rows = sheet_table(wb, INFLCAT_SHEET)
    require(c_h, [HEADLINE_COL], INFLCAT_SHEET)
    headline = col(c_rows, c_h, HEADLINE_COL)

    # ---- the education K -------------------------------------------------
    u_h, u_rows = sheet_table(wb, URATE_SHEET)
    require(u_h, [URATE_COL], URATE_SHEET)
    urate_gap = col(u_rows, u_h, URATE_COL)

    wb.close()

    # ---- derived ---------------------------------------------------------
    ratio_low  = ratio_3m(retail["low"],  retail["mid"])
    ratio_high = ratio_3m(retail["high"], retail["mid"])
    yoy = {k: yoy_from_index(v) for k, v in retail.items()}

    # Group inflation = headline rate + that group's published gap.
    hmap = dict(headline)
    infl_level = {}
    for key, gaps in infl.items():
        infl_level[key.replace("infl_", "rate_")] = [
            (m, None if (g is None or hmap.get(m) is None) else hmap[m] + g)
            for m, g in gaps
        ]

    latest_month, _ = latest(retail["mid"])

    # ---- CSV baselines ---------------------------------------------------
    months = [m for m, _ in retail["mid"]]
    idx = lambda s: dict(s)
    spend_rows = [
        [m,
         idx(retail["low"]).get(m),  idx(retail["mid"]).get(m),  idx(retail["high"]).get(m),
         idx(foodbev["low"]).get(m), idx(foodbev["mid"]).get(m), idx(foodbev["high"]).get(m),
         idx(tops["top_125_175"]).get(m), idx(tops["top_175_225"]).get(m),
         idx(tops["top_225_250"]).get(m), idx(tops["top_250p"]).get(m)]
        for m in months
    ]
    spend_changed = write_csv(
        SPEND_CSV,
        ["month", "low_retail", "mid_retail", "high_retail",
         "low_foodbev", "mid_foodbev", "high_foodbev",
         "top_125_175", "top_175_225", "top_225_250", "top_250p"],
        spend_rows)

    gmonths = [m for m, _ in infl["infl_top20"]]
    gaps_rows = [
        [m, idx(headline).get(m),
         idx(infl["infl_bottom40"]).get(m), idx(infl["infl_mid40"]).get(m),
         idx(infl["infl_top20"]).get(m),
         idx(urate_gap).get(m)]
        for m in gmonths
    ]
    gaps_changed = write_csv(
        GAPS_CSV,
        ["month", "infl_headline", "infl_bottom40", "infl_mid40", "infl_top20",
         "edu_urate_gap"],
        gaps_rows)

    # ---- payload ---------------------------------------------------------
    payload = {
        "build_time": dt.datetime.utcnow().isoformat() + "Z",
        "latest_label": latest_month,
        "source": "Federal Reserve Bank of New York, Economic Heterogeneity Indicators",
        "kpis": {
            "spend_low_yoy":   kpi(yoy["low"]),
            "spend_mid_yoy":   kpi(yoy["mid"]),
            "spend_high_yoy":  kpi(yoy["high"]),
            "ratio_high_mid":  kpi(ratio_high, decimals=3),
            "ratio_low_mid":   kpi(ratio_low,  decimals=3),
            "infl_gap":        kpi([(m, (b or 0) - (t or 0)) if (b is not None and t is not None)
                                    else (m, None)
                                    for (m, b), (_, t) in zip(infl["infl_bottom40"],
                                                              infl["infl_top20"])]),
        },
        "spend_ratio_low":  to_pairs(ratio_low, 4),
        "spend_ratio_high": to_pairs(ratio_high, 4),
        "spend_low":  to_pairs(retail["low"]),
        "spend_mid":  to_pairs(retail["mid"]),
        "spend_high": to_pairs(retail["high"]),
        "spend_yoy_low":  to_pairs(yoy["low"], 2),
        "spend_yoy_mid":  to_pairs(yoy["mid"], 2),
        "spend_yoy_high": to_pairs(yoy["high"], 2),
        "top_125_175": to_pairs(tops["top_125_175"]),
        "top_175_225": to_pairs(tops["top_175_225"]),
        "top_225_250": to_pairs(tops["top_225_250"]),
        "top_250p":    to_pairs(tops["top_250p"]),
        "foodbev_low":  to_pairs(foodbev["low"]),
        "foodbev_mid":  to_pairs(foodbev["mid"]),
        "foodbev_high": to_pairs(foodbev["high"]),
        "infl_headline": to_pairs(headline, 2),
        "infl_bottom40": to_pairs(infl["infl_bottom40"], 2),
        "infl_mid40":    to_pairs(infl["infl_mid40"], 2),
        "infl_top20":    to_pairs(infl["infl_top20"], 2),
        "rate_bottom40": to_pairs(infl_level["rate_bottom40"], 2),
        "rate_mid40":    to_pairs(infl_level["rate_mid40"], 2),
        "rate_top20":    to_pairs(infl_level["rate_top20"], 2),
        "edu_urate_gap": to_pairs(urate_gap, 2),
    }

    payload["notice"] = (
        "The NY Fed publishes these indicators quarterly (February, May and "
        "September), so this page updates three times a year and runs a few "
        "months behind the monthly data it contains. Latest observation: "
        "%s. The EHIs are not official estimates of the Federal Reserve."
        % (latest_month or "n/a"))

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUT_PATH.open("w") as f:
        json.dump(payload, f, separators=(",", ":"))

    print("\nWrote %s" % OUT_PATH, flush=True)
    print("  latest_label:     %s" % latest_month, flush=True)
    print("  spending obs:     %d months (%s .. %s)"
          % (len(months), months[0] if months else "-", months[-1] if months else "-"), flush=True)
    print("  gaps obs:         %d months" % len(gmonths), flush=True)
    print("  low  YoY:         %s" % payload["kpis"]["spend_low_yoy"], flush=True)
    print("  mid  YoY:         %s" % payload["kpis"]["spend_mid_yoy"], flush=True)
    print("  high YoY:         %s" % payload["kpis"]["spend_high_yoy"], flush=True)
    print("  high/mid ratio:   %s" % payload["kpis"]["ratio_high_mid"], flush=True)
    print("  low/mid  ratio:   %s" % payload["kpis"]["ratio_low_mid"], flush=True)
    print("  CSV changed:      spending=%s gaps=%s"
          % (spend_changed, gaps_changed), flush=True)


if __name__ == "__main__":
    try:
        main()
    except (error.URLError, error.HTTPError, RuntimeError, ValueError) as e:
        # Non-blocking by design: leave the committed baselines in place.
        print("FETCH FAILED (leaving existing data untouched): %s" % e,
              file=sys.stderr)
        sys.exit(0)
